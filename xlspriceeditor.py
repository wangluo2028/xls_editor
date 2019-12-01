#pyinstaller --onefile xlspriceeditor.py
import os
#from openpyxl import Workbook, load_workbook
import openpyxl
#import openpyxl.utils
#from openpyxl.utils import rows_from_range
import tkinter as tk
from tkinter import filedialog
import sheettable
import notepad
import re

def string_conv2_number(number_string):
    try:
        return int(number_string)
    except ValueError:
        return float(number_string)

def price_modify(price):
    if (price < 600):
        price += 50
    elif (price < 1000):
        price += 150
    elif (price < 2000):
        price += 200
    else:
        price += 250

    return price

def number_handle(number):
    number_string = number.group()
    number = string_conv2_number(number_string)
    number = price_modify(number)
    number_string = str(number)
    return number_string

def price_auto_modify(cell):
    value = cell.value
    if isinstance(value, (int, float, complex)):
        value = price_modify(value)
        cell.value = value
    elif isinstance(value, str):
        re_pattern = re.compile('\d+', re.S)
        new_value = re_pattern.sub(number_handle, value)
        cell.value = new_value

def check_merged_cell(sheet, cell):
    is_merged_cell = False
    cell_coord = cell.coordinate
    cell_column = cell.column
    cell_row = cell.row
    merged_cell_ranges = sheet.merged_cell_ranges
    for merged_cell_range in merged_cell_ranges:
        if merged_cell_range.min_col <= cell_column and cell_column <= merged_cell_range.max_col\
            and merged_cell_range.min_row <= cell_row and cell_row <= merged_cell_range.max_row:
            is_merged_cell = True
            break

    return is_merged_cell

def getExcel():
    # root.withdraw()
    xls_file = filedialog.askopenfilename(title="请选择Excel文件",
                                          filetypes=(('xlsx files', '*.xlsx'), ('xls files', '*.xls')))
    if not os.path.exists(xls_file):
        print('file not exist')
        return
    # xls_file = 'D:/User/ProgramData/wechat/WeChat Files/bstuder23/FileStorage/File/2019-12/江西新凯勤12月1号全系报价具体价格以B2B为准(1).xlsx'
    workbook_from_file = openpyxl.load_workbook(xls_file)
    # sheet['B4'].value = sheet['B4'].value - 100
    for sheet in workbook_from_file.__iter__():
        merged_cell = sheet.merged_cell_ranges
        for row in sheet.iter_rows():
            for cell in row:
                # check is price cell
                # if (cell.col_idx % 2 == 0):
                if (cell.column % 2 == 0):
                    if (not check_merged_cell(sheet, cell)):
                        price_auto_modify(cell)

    file_dir = os.path.dirname(xls_file)
    file_name = os.path.basename(xls_file)
    base_file_name, file_ext = os.path.splitext(file_name)
    new_file_name = file_dir + '/' + base_file_name + '_新' + file_ext
    workbook_from_file.save(new_file_name)

def main():
    app = sheettable.SheetTable()
    app.mainloop()

    notepad = notepad.Notepad(width=600, height=400)
    notepad.run()

    root = tk.Tk()
    canvas1 = tk.Canvas(root, width=300, height=300, bg='lightsteelblue')
    canvas1.pack()
    browseButton_Excel = tk.Button(text='请打开Excel文件', command=getExcel, bg='green', fg='white',
                                   font=('helvetica', 12, 'bold'))
    canvas1.create_window(150, 150, window=browseButton_Excel)
    root.mainloop()

if __name__ == "__main__":
    main()